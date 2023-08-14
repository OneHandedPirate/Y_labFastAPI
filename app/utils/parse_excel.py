import openpyxl


def excel_to_list(file_path: str) -> list:
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    worksheet = workbook.active

    menus: list = []

    current_menu: dict = {}
    current_submenu: dict = {}

    for row in worksheet.iter_rows(min_row=1, values_only=True):
        if row[0]:
            if current_submenu:
                current_menu['submenus'].append(current_submenu)
            current_submenu = {}
            if current_menu:
                menus.append(current_menu)
            current_menu = {'id': str(row[0]), 'title': row[1], 'description': row[2], 'submenus': []}
        elif type(row[1]) == int:
            if current_menu and current_submenu:
                current_menu['submenus'].append(current_submenu)
            current_submenu = {'id': str(row[1]), 'title': row[2], 'description': row[3], 'dishes': []}
        elif type(row[2]) == int:
            current_dish = {'id': str(row[2]), 'title': row[3], 'description': row[4], 'price': str(row[5])}
            current_submenu['dishes'].append(current_dish)

    if current_menu:
        if current_submenu:
            current_menu['submenus'].append(current_submenu)
        menus.append(current_menu)

    return menus


def excel_to_tuple(file_path: str) -> tuple:
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    worksheet = workbook.active

    menus: list = []
    submenus: list = []
    dishes: list = []

    for row in worksheet.iter_rows(min_row=1, values_only=True):
        if row[0]:
            menus.append({'id': str(row[0]), 'title': row[1], 'description': row[2]})
        elif type(row[1]) == int:
            submenus.append({'menu_id': menus[-1]['id'], 'id': str(row[1]), 'title': row[2], 'description': row[3]})
        elif type(row[2]) == int:
            dishes.append({'menu_id': menus[-1]['id'], 'submenu_id': submenus[-1]['id'],
                           'id': str(row[2]), 'title': row[3],
                           'description': row[4], 'price': str(row[5])})
    return menus, submenus, dishes
